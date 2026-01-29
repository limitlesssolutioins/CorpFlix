import pandas as pd
import openpyxl

# Leer archivo Excel
file_path = r'C:\Users\USUARIO\Desktop\Limitless Solutions\Lidus\src\documentacion\Riesgos.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

output = []

for sheet_name in wb.sheetnames:
    output.append("=" * 100)
    output.append(f"HOJA: {sheet_name}")
    output.append("=" * 100)
    output.append("")
    
    ws = wb[sheet_name]
    
    # Obtener todas las celdas con valores
    max_row = ws.max_row
    max_col = ws.max_column
    
    output.append(f"Dimensiones: {max_row} filas x {max_col} columnas")
    output.append("")
    
    # Mostrar contenido fila por fila (primeras 30 filas)
    for row_idx in range(1, min(31, max_row + 1)):
        row_data = []
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None and str(cell.value).strip() != "":
                row_data.append(f"Col{col_idx}: {cell.value}")
        
        if row_data:
            output.append(f"Fila {row_idx}: {' | '.join(row_data)}")
    
    output.append("")
    output.append("")

# Guardar a archivo
with open('analisis_riesgos.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(output))

print("Análisis guardado en: analisis_riesgos.txt")
print(f"Total de líneas: {len(output)}")
